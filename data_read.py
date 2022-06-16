from openpyxl import load_workbook

class ExploreData():
    def __init__(self):
        self.filename="data.xlsx"
        self.data=[]

    def open_file(self):
        self.wb=load_workbook(self.filename)
        self.ws=self.wb['Sayfa1']
        self.data_area_in_file=self.ws['b1:M37']


    def create_data_json(self):
        temp_data = {}
        data_index = 0

        for row_index, row_values in enumerate(self.data_area_in_file):
            if row_index == 0:
                for year in range(2005, 2023, 1):
                    temp_data["year"] = year
                    temp_data["month"] = []
                    for cell_index, cell in enumerate(row_values):
                        temp_data["month"].append({
                            "month_number": cell_index + 1,
                            "month_name": cell.value,
                            "year_inflation_value": "",
                            "month_inflation_value": ""
                        })
                    self.data.append(temp_data)
                    temp_data = {}
            elif row_index % 2 == 1:
                for index, month in enumerate(self.data[data_index]['month']):
                    month['year_inflation_value'] = row_values[index].value
            elif row_index % 2 == 0:
                for index, month in enumerate(self.data[data_index]['month']):
                    month['month_inflation_value'] = row_values[index].value
                data_index += 1

    def save_file(self):
        # self.wb.close()
        self.wb.save(self.filename)

    def close_file(self):
        self.wb.close()

    def get_data(self):
        self.open_file()
        self.create_data_json()
        # self.save_file()
        # self.close_file()
        return self.data


# file = "data.xlsx"
# wb = load_workbook(file)
#
# ws = wb["Sayfa1"]
# data = []
# temp_data = {}
# data_index=0
# for row_index,row_values in enumerate(ws["b1:M31"]):
#     # print(row_values[0].value)
#     if row_index==0:
#         for year in range(2008,2024,1):
#             temp_data["year"]=year
#             temp_data["month"]=[]
#             for cell_index,cell in enumerate(row_values):
#                 temp_data["month"].append({
#                 "month_number":cell_index+1,
#                 "month_name":cell.value,
#                 "year_inflation_value": "",
#                 "month_inflation_value":""
#                 })
#             data.append(temp_data)
#             temp_data={}
#     elif row_index%2==1:
#         for index,month in enumerate(data[data_index]['month']):
#             month['year_inflation_value']=row_values[index].value
#     elif row_index%2==0:
#         for index,month in enumerate(data[data_index]['month']):
#             month['month_inflation_value'] = row_values[index].value
#         data_index +=1
#
# wb.save(file)

if __name__=='__main__':
    # filename="data.xlsx"
    infilation_data=ExploreData()
    print(infilation_data.get_data())
    # get_all_infilation_data=infilation_data.get_data()
    #
    # print(get_all_infilation_data)
